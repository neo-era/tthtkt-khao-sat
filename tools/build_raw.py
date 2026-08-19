# -*- coding: utf-8 -*-
"""
Sinh mảng RAW (danh mục vị trí khảo sát) từ file Excel nguồn và ghi đè vào index.html.

    pip install openpyxl
    python tools/build_raw.py

Nguồn : docs/DS TRẠM DỪNG XE BUÝT.xlsx
    cột 1 STT (công thức, bỏ qua – script đánh số lại)
    cột 2 Tuyến đường | cột 3 Số nhà | cột 4 Phường/Xã | cột 5 Hạ tầng
    cột 6 mã địa bàn (Q1/Q3/Q5/Q8/Q10/Q11/PN/BT) – CHỈ có ở các quận thuộc phạm vi khảo sát.

Phạm vi: chỉ lấy các dòng CÓ mã ở cột 6 (610 vị trí thuộc 8 quận). Các dòng còn lại
(khu vực TP.HCM ngoài 8 quận, Vũng Tàu, Bình Dương) không thuộc phạm vi, bị bỏ qua.

Đích  : dòng `const RAW = [...];` trong index.html (thay nguyên dòng).

⚠ Script này GHI ĐÈ toàn bộ mảng RAW, nên nó xóa luôn 51 nhà chờ của công văn 2858.
Chạy xong phải chạy tiếp `python tools/them_nha_cho_2858.py` để chèn lại, nếu không
danh mục tụt từ 661 xuống 610 vị trí.

Danh sách gốc không có cột Mã điểm dừng nên 610 vị trí này để `ma` rỗng; chỉ nhà chờ
lấy từ công văn 2858 mới có mã.
"""
import io
import json
import os
import re
import unicodedata
from collections import Counter

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "docs", "DS TRẠM DỪNG XE BUÝT.xlsx")
HTML = os.path.join(ROOT, "index.html")

# Mã địa bàn ở cột 6 -> tên hiển thị. Thứ tự dict cũng là thứ tự sắp xếp/chip lọc.
DIABAN = {
    "Q1": "Quận 1", "Q3": "Quận 3", "Q5": "Quận 5", "Q8": "Quận 8",
    "Q10": "Quận 10", "Q11": "Quận 11", "PN": "Phú Nhuận", "BT": "Bình Thạnh",
}

# Dấu thanh kiểu cũ -> kiểu mới, để "Hoà Hưng" và "Hòa Hưng" không tách thành 2 phường.
TONE = [("oà", "òa"), ("oá", "óa"), ("oả", "ỏa"), ("oã", "õa"), ("oạ", "ọa"),
        ("oè", "òe"), ("oé", "óe"), ("oẻ", "ỏe"), ("uỳ", "ùy"), ("uý", "úy")]


def cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(v))).strip()


def norm_loai(s):
    """Thống nhất cách viết: 'Loại' -> 'loại', '4s' -> '4S'."""
    s = re.sub(r"\bLoại\b", "loại", s)
    s = re.sub(r"loại\s*(\d)\s*[sS]\b", lambda m: "loại " + m.group(1) + "S", s)
    return s


def norm_phuong(s):
    s = re.sub(r"^(Phường|Xã|Thị trấn)\s+", "", s).strip()
    for a, b in TONE:
        s = s.replace(a, b)
    return s


def main():
    ws = openpyxl.load_workbook(XLSX, data_only=True).active

    raw = []
    skipped = 0
    for r in range(5, ws.max_row + 1):
        duong = cell(ws.cell(r, 2).value)
        if not duong:
            continue                       # dòng tiêu đề khu vực (I / II / III)
        ma = cell(ws.cell(r, 6).value)
        if ma not in DIABAN:
            skipped += 1                   # ngoài phạm vi 8 quận
            continue
        sonha = cell(ws.cell(r, 3).value)
        raw.append({
            "stt": 0,
            "ma": "",                      # danh sách gốc không có Mã điểm dừng
            "diaban": DIABAN[ma],
            "ten": duong + (" – " + sonha if sonha else ""),
            "duong": duong,
            "sonha": sonha,
            "phuong": norm_phuong(cell(ws.cell(r, 4).value)),
            "loai": norm_loai(cell(ws.cell(r, 5).value)),
            "ghichu": "",
        })

    # Gom liền mạch theo địa bàn -> phường/xã (giữ thứ tự gốc trong từng phường), rồi đánh lại STT.
    order = list(DIABAN.values())
    raw.sort(key=lambda x: (order.index(x["diaban"]), x["phuong"]))
    for i, x in enumerate(raw):
        x["stt"] = i + 1

    print("Số vị trí   :", len(raw), "(bỏ qua", skipped, "vị trí ngoài phạm vi)")
    print("Địa bàn     :", dict(Counter(x["diaban"] for x in raw)))
    print("Phường/xã   :", len(set(x["phuong"] for x in raw)))
    print("Hạ tầng     :", dict(Counter(x["loai"] for x in raw)))

    line = "const RAW = " + json.dumps(raw, ensure_ascii=False) + ";"
    with io.open(HTML, encoding="utf-8") as f:
        src = f.read()
    new, n = re.subn(r"(?m)^const RAW = \[.*\];$", lambda m: line, src, count=1)
    assert n == 1, "Không tìm thấy dòng 'const RAW = [...];' trong index.html"
    with io.open(HTML, "w", encoding="utf-8", newline="") as f:
        f.write(new)
    print("Đã ghi", HTML)


if __name__ == "__main__":
    main()
