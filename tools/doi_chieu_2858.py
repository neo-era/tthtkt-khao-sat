# -*- coding: utf-8 -*-
"""
Đối chiếu danh sách nhà chờ xe buýt trong công văn 2858 với 610 trụ dừng đang khảo sát.

    pip install openpyxl
    python tools/doi_chieu_2858.py

Nguồn : 2858 KTHT DS gửi TTHTKT.xlsx (gốc repo)
          sheet DS1              -> 19 nhà chờ tăng mảng xanh (nhóm A+B+C, bỏ nhóm D Thủ Đức)
          sheet DANH SACH 75VITRI -> 39 nhà chờ cần chiếu sáng thuộc Chiếu sáng khu vực Trung Tâm
        mảng RAW trong index.html -> 610 trụ dừng đang khảo sát (Q1/Q3/Q5/Q8/Q10/Q11/PN/BT)
Đích  : docs/Doi_chieu_2858_v1.0.xlsx

LƯU Ý VỀ BẢN CHẤT DỮ LIỆU: hai danh mục đếm hai loại hạ tầng KHÁC NHAU. 610 vị trí
đang khảo sát là 100% *trụ dừng*, còn công văn 2858 là 100% *nhà chờ*. Nên bảng này
không phải để tìm trùng lặp (gần như không có), mà để biết mỗi nhà chờ nằm ở đâu so
với các trụ dừng đang khảo sát — cùng đoạn đường thì đi khảo sát chung một chuyến.

Script chỉ ĐỌC index.html, không sửa gì trong app.
"""
import collections
import io
import json
import os
import re
import unicodedata

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Dùng lại logic chuẩn hóa của build_raw.py, đừng viết lần hai.
from build_raw import TONE, cell, norm_phuong

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONGVAN = os.path.join(ROOT, "2858 KTHT DS gửi TTHTKT.xlsx")
HTML = os.path.join(ROOT, "index.html")
DICH = os.path.join(ROOT, "docs", "Doi_chieu_2858_v1.0.xlsx")

DON_VI = "Chiếu sáng khu vực Trung tâm"

# Mã điểm dừng trong công văn -> tên địa bàn (để thống kê theo quận).
QUAN = {
    "Q1": "Quận 1", "Q3": "Quận 3", "Q5": "Quận 5", "Q8": "Quận 8",
    "Q10": "Quận 10", "Q11": "Quận 11", "QPN": "Phú Nhuận", "QBTH": "Bình Thạnh",
}

# Mã trạm: "Q3 101", "QBTH 149", "TPTĐ 225". Viết [A-Z]+\d+ sẽ chỉ bắt được "Q3".
RE_MA = re.compile(r"M[ãa]\s*[Tt]r[ạa]m\s*:?\s*([A-ZĐ]{1,4}\d{0,2}\s+\d{1,3})")

MUC_TRUNG = "Trùng đúng địa chỉ"
MUC_SONHA = "Cùng số nhà"
MUC_DOAN = "Cùng đoạn đường"
MUC_KHACPHUONG = "Cùng đường, khác phường"
MUC_KHONG = "Không có trụ dừng"

MAU = {
    MUC_TRUNG: "C6E7C6",
    MUC_SONHA: "E2F0D9",
    MUC_DOAN: "DEEAF6",
    MUC_KHACPHUONG: "FFE699",
    MUC_KHONG: "FBD5B5",
}


# ------------------------------- chuẩn hóa -------------------------------

def ch(s):
    """Chuẩn hóa để so khớp: NFC, gộp khoảng trắng, thường hóa, thống nhất dấu thanh."""
    s = cell(s).lower()
    for a, b in TONE:
        s = s.replace(a, b)
    return s


def nph(s):
    """Phường/xã đã bỏ tiền tố, dùng làm khóa so khớp."""
    return ch(norm_phuong(cell(s)))


def tach_sonha(s):
    """('đối diện 115 - 117 (X)') -> (True, '115', '115 - 117').

    Trả về (đối_diện, số_đầu_tiên, chuỗi_đã_gọn). Phần trong ngoặc bị bỏ vì hai
    nguồn ghi chú khác nhau: '252 (Đối diện 175)' và '252' là cùng một chỗ.
    """
    t = ch(s)
    doi_dien = bool(re.match(r"^(đối diện|đd)\b", t))
    t = re.sub(r"^(đối diện|đd)\s*(số\s*)?", "", t)
    t = re.sub(r"\(.*?\)", "", t).strip()
    m = re.search(r"\d+", t)
    return doi_dien, (m.group(0) if m else ""), t


def quan_cua(ma):
    return QUAN.get(re.split(r"\s+", ma)[0] if ma else "", "")


# ------------------------------- đọc nguồn -------------------------------

def doc_tru_dung():
    """610 trụ dừng từ mảng RAW trong index.html (chỉ đọc, không sửa)."""
    with io.open(HTML, encoding="utf-8") as f:
        src = f.read()
    m = re.search(r"^const RAW = (\[.*\]);$", src, re.M)
    assert m, "Không tìm thấy dòng 'const RAW = [...];' trong index.html"
    return json.loads(m.group(1))


def doc_39(wb):
    """39 nhà chờ: thuộc Chiếu sáng khu vực Trung Tâm VÀ có đánh dấu chiếu sáng."""
    ws = wb["DANH SACH 75VITRI"]
    tat_ca, ds = {}, []
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 11)]
        ma = cell(v[1])
        if not ma:
            continue
        # Dòng cuối sheet là dòng TỔNG CỘNG (mảng xanh 186 / chiếu sáng 75), không
        # phải một vị trí. Nó không có mã điểm dừng nên đã bị loại ở trên.
        rec = {
            "ma": ma, "ten": cell(v[2]), "sonha": cell(v[3]), "duong": cell(v[4]),
            "phuong": cell(v[5]), "loai": cell(v[6]),
        }
        tat_ca[ma] = rec
        if cell(v[9]) == DON_VI and cell(v[8]):
            ds.append(rec)
    return ds, tat_ca


def doc_19(wb, tat_ca):
    """19 nhà chờ mảng xanh ở sheet DS1 — nhóm A + B + C, bỏ nhóm D (Thủ Đức).

    Sheet này cấu trúc không đều, ba chỗ phải xử lý riêng:
      - một ô 'Địa chỉ' chứa NHIỀU vị trí, ngăn bằng dấu '+';
      - dòng tiêu đề nhóm đôi khi tự nó mang địa chỉ (nhóm C ghi tên đường ngay ở
        ô 'C. Đường Nguyễn Hữu Cảnh', cột 'Tuyến đường' để trống) — chỉ đọc dòng có
        cột 'Tuyến đường' thì ra 18 chứ không phải 19;
      - hai vị trí Điện Biên Phủ không có mã trạm, vẫn phải giữ.
    """
    ws = wb["DS1"]
    ds, nhom = [], ""
    for r in range(6, ws.max_row + 1):
        c1 = cell(ws.cell(r, 1).value)
        if re.match(r"^[A-D]\.", c1):
            nhom = c1
        if nhom.startswith("D"):
            continue                      # Xa lộ Hà Nội: Q2/QTD/TPTĐ, ngoài 8 quận
        dia_chi = ws.cell(r, 6).value
        if not dia_chi:
            continue
        duong_row = cell(ws.cell(r, 2).value)
        duong = duong_row or re.sub(r"^[A-D]\.\s*(Đường|Xa lộ)?\s*", "", c1).strip()
        sau = cell(ws.cell(r, 3).value)
        hai_ben = cell(ws.cell(r, 4).value)
        for phan in str(dia_chi).split("+"):
            phan = cell(phan).rstrip(";").strip()
            if not phan or phan.startswith("(vị trí"):
                continue                  # ghi chú phạm vi của cả nhóm
            m = RE_MA.search(phan)
            ma = re.sub(r"\s+", " ", m.group(1)) if m else ""
            mo_ta = re.sub(r"\s*\(M[ãa].*", "", phan).strip()
            s2 = tat_ca.get(ma)
            ds.append({
                "nhom": nhom[:1], "ma": ma, "mo_ta": mo_ta,
                "duong": s2["duong"] if s2 else duong,
                "sonha": s2["sonha"] if s2 else mo_ta,
                "phuong": s2["phuong"] if s2 else "",
                "loai": s2["loai"] if s2 else "",
                "ten": s2["ten"] if s2 else "",
                "sau": sau, "hai_ben": hai_ben, "co_s2": bool(s2),
            })
    return ds


# ------------------------------- đối chiếu -------------------------------

def lap_chi_muc(tru_dung):
    by_dp, by_d = collections.defaultdict(list), collections.defaultdict(list)
    for t in tru_dung:
        by_dp[(ch(t["duong"]), nph(t["phuong"]))].append(t)
        by_d[ch(t["duong"])].append(t)
    return by_dp, by_d


def doi_chieu(x, by_dp, by_d):
    """Trả (mức khớp, danh sách trụ dừng liên quan).

    Khóa là (đường, phường) chứ không phải mỗi tên đường: danh sách gốc có 43 vị trí
    'Lê Lợi' nhưng 32 ở Vũng Tàu và 11 ở Hóc Môn — khớp theo tên đường thôi thì
    nhà chờ Lê Lợi ở Quận 1 sẽ bị gán nhầm vào trụ dừng ngoài tỉnh.
    """
    d, p = ch(x["duong"]), nph(x["phuong"])
    ung = by_dp.get((d, p), [])
    if not ung and not p:
        ung = by_d.get(d, [])             # công văn thiếu phường: đành xét cả đường
    if ung:
        dd, so, gon = tach_sonha(x["sonha"])
        for t in ung:
            dd2, so2, gon2 = tach_sonha(t["sonha"])
            if gon and gon == gon2 and dd == dd2:
                return MUC_TRUNG, [t]
        for t in ung:
            _, so2, _ = tach_sonha(t["sonha"])
            if so and so == so2:
                return MUC_SONHA, [t]
        return MUC_DOAN, ung[:5]
    if by_d.get(d):
        return MUC_KHACPHUONG, by_d[d][:5]
    return MUC_KHONG, []


def mo_ta_ung(ds):
    return " ; ".join("STT %d – %s" % (t["stt"], t["sonha"] or "(không số)") for t in ds)


def ghi_chu(x, muc):
    """Những chỗ máy không quyết được, phải nói rõ để người đọc tự xác nhận."""
    n = []
    if muc == MUC_KHACPHUONG:
        n.append("Phường ghi khác danh mục 610 — cần xác nhận có cùng đoạn đường không")
    if muc == MUC_SONHA:
        n.append("Số nhà trùng nhưng cách ghi khác — cần xác nhận")
    if not x.get("ma"):
        n.append("Công văn không ghi mã trạm")
    if "co_s2" in x and not x["co_s2"] and x.get("ma"):
        n.append("Mã trạm không có trong sheet 186 vị trí — thiếu số nhà, phường")
    # Hai sheet trong cùng công văn có chỗ ghi lệch nhau (vd Q3 027: DS1 ghi
    # "Đối diện số 208", sheet 186 vị trí ghi "Đối diện số 206 (247)").
    if x.get("co_s2") and x.get("mo_ta"):
        a = re.search(r"\d+", ch(x["mo_ta"]))
        b = re.search(r"\d+", ch(x["sonha"]))
        if a and b and a.group(0) != b.group(0):
            n.append('DS1 ghi "%s", sheet 186 vị trí ghi "%s"' % (x["mo_ta"], x["sonha"]))
    return " | ".join(n)


# ------------------------------- ghi Excel -------------------------------

FONT = "Times New Roman"
VIEN = Border(*[Side(style="thin", color="FFC9D4DF")] * 4)


def to_sheet(ws, tieu_de, rong, dong, cot_muc=None):
    ws.append(tieu_de)
    for i, c in enumerate(ws[1], start=1):
        c.font = Font(name=FONT, bold=True, color="FFFFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="FF0F2A43")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = VIEN
        ws.column_dimensions[get_column_letter(i)].width = rong[i - 1]
    ws.row_dimensions[1].height = 34

    for hang in dong:
        ws.append(hang)
    for r in range(2, ws.max_row + 1):
        mau = MAU.get(ws.cell(r, cot_muc).value) if cot_muc else None
        for c in range(1, len(tieu_de) + 1):
            o = ws.cell(r, c)
            o.font = Font(name=FONT, size=11)
            o.alignment = Alignment(vertical="top", wrap_text=True)
            o.border = VIEN
            if mau:
                o.fill = PatternFill("solid", fgColor="FF" + mau)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = "A1:%s1" % get_column_letter(len(tieu_de))


def main():
    wb_cv = openpyxl.load_workbook(CONGVAN, data_only=True)
    tru_dung = doc_tru_dung()
    ds39, tat_ca = doc_39(wb_cv)
    ds19 = doc_19(wb_cv, tat_ca)
    by_dp, by_d = lap_chi_muc(tru_dung)

    for x in ds39 + ds19:
        x["muc"], x["ung"] = doi_chieu(x, by_dp, by_d)

    ma39 = {x["ma"] for x in ds39 if x["ma"]}
    ma19 = {x["ma"] for x in ds19 if x["ma"]}
    chung = sorted(ma39 & ma19)

    out = openpyxl.Workbook()
    out.remove(out.active)

    # --- sheet 1: 39 nhà chờ chiếu sáng ---
    ws = out.create_sheet("39 nhà chờ chiếu sáng")
    to_sheet(ws,
        ["TT", "Mã điểm dừng", "Tên điểm dừng", "Số nhà", "Đường", "Phường",
         "Loại nhà chờ", "Mức khớp", "Trụ dừng đang khảo sát trên cùng đoạn",
         "Có trong danh sách 19", "Ghi chú"],
        [5, 13, 26, 22, 20, 16, 18, 20, 40, 13, 40],
        [[i, x["ma"], x["ten"], x["sonha"], x["duong"], x["phuong"], x["loai"],
          x["muc"], mo_ta_ung(x["ung"]) or "—",
          "Có" if x["ma"] in ma19 else "", ghi_chu(x, x["muc"])]
         for i, x in enumerate(ds39, 1)],
        cot_muc=8)

    # --- sheet 2: 19 nhà chờ mảng xanh ---
    ws = out.create_sheet("19 nhà chờ mảng xanh")
    to_sheet(ws,
        ["TT", "Nhóm", "Mã điểm dừng", "Mô tả trong công văn", "Số nhà", "Đường",
         "Phường", "Loại nhà chờ", "Trồng phía sau (cả tuyến)", "Trồng 2 bên (cả tuyến)",
         "Mức khớp", "Trụ dừng đang khảo sát trên cùng đoạn", "Có trong danh sách 39",
         "Ghi chú"],
        [5, 7, 13, 30, 20, 20, 16, 18, 12, 12, 20, 40, 13, 40],
        [[i, x["nhom"], x["ma"], x["mo_ta"], x["sonha"], x["duong"], x["phuong"],
          x["loai"], x["sau"], x["hai_ben"], x["muc"], mo_ta_ung(x["ung"]) or "—",
          "Có" if x["ma"] and x["ma"] in ma39 else "", ghi_chu(x, x["muc"])]
         for i, x in enumerate(ds19, 1)],
        cot_muc=11)

    # --- sheet 3: gộp theo tuyến đường (dùng để xếp lịch đi hiện trường) ---
    gop = collections.OrderedDict()
    for x in ds39 + ds19:
        if x["ma"] and x["ma"] in chung and x in ds19:
            continue                      # đã tính ở danh sách 39, không đếm hai lần
        k = (x["duong"], x["phuong"])
        g = gop.setdefault(k, {"n": 0, "ma": [], "ds": set(), "ung": [], "muc": set()})
        g["n"] += 1
        g["ma"].append(x["ma"] or "(không mã)")
        g["ds"].add("39" if x in ds39 else "19")
        g["muc"].add(x["muc"])
        if not g["ung"]:
            g["ung"] = x["ung"]
    ws = out.create_sheet("Gộp theo tuyến đường")
    to_sheet(ws,
        ["Đường", "Phường", "Số nhà chờ", "Mã điểm dừng", "Thuộc danh sách",
         "Trụ dừng đang khảo sát trên cùng đoạn"],
        [24, 18, 10, 34, 14, 46],
        [[k[0], k[1], g["n"], ", ".join(g["ma"]),
          " + ".join(sorted(g["ds"])), mo_ta_ung(g["ung"]) or "—"]
         for k, g in gop.items()])

    # --- sheet 4: tổng hợp ---
    d39 = collections.Counter(x["muc"] for x in ds39)
    d19 = collections.Counter(x["muc"] for x in ds19)
    q39 = collections.Counter(quan_cua(x["ma"]) or "(không rõ)" for x in ds39)
    q19 = collections.Counter(quan_cua(x["ma"]) or "(không rõ)" for x in ds19)
    thu_tu = [MUC_TRUNG, MUC_SONHA, MUC_DOAN, MUC_KHACPHUONG, MUC_KHONG]
    dong = [["BẢN CHẤT DỮ LIỆU", "", ""],
            ["610 vị trí đang khảo sát", "100% trụ dừng (loại 1/2/3/4/4S/5/5S)", ""],
            ["Công văn 2858", "100% nhà chờ (6180, 3090, 1876S, 2666, 3299, N03…)", ""],
            ["", "Hai loại hạ tầng khác nhau nên gần như không có vị trí trùng.", ""],
            ["", "", ""],
            ["MỨC KHỚP", "Danh sách 39", "Danh sách 19"]]
    dong += [[m, d39.get(m, 0), d19.get(m, 0)] for m in thu_tu]
    dong += [["Tổng", len(ds39), len(ds19)],
             ["", "", ""],
             ["Trùng giữa hai danh sách", len(chung), ", ".join(chung)],
             ["Tổng vị trí riêng biệt", len(ds39) + len(ds19) - len(chung), ""],
             ["", "", ""],
             ["THEO ĐỊA BÀN", "Danh sách 39", "Danh sách 19"]]
    dong += [[q, q39.get(q, 0), q19.get(q, 0)]
             for q in sorted(set(q39) | set(q19))]
    dong += [["", "", ""], ["CẦN XÁC NHẬN BẰNG MẮT", "", ""]]
    for x in ds39 + ds19:
        g = ghi_chu(x, x["muc"])
        if g:
            dong.append([x["ma"] or "(không mã)", "%s %s" % (x["duong"], x["sonha"]), g])
    ws = out.create_sheet("Tổng hợp")
    to_sheet(ws, ["Chỉ tiêu", "Giá trị", "Ghi chú"], [34, 30, 60], dong)
    ws.auto_filter.ref = None             # sheet này là bảng tổng hợp, không lọc

    out.save(DICH)

    print("Trụ dừng đang khảo sát :", len(tru_dung))
    print("Danh sách 39           :", len(ds39), dict(d39))
    print("Danh sách 19           :", len(ds19),
          dict(collections.Counter(x["nhom"] for x in ds19)), dict(d19))
    print("Trùng hai danh sách    :", len(chung), chung)
    print("Vị trí riêng biệt      :", len(ds39) + len(ds19) - len(chung))
    print("Đã ghi", DICH)


if __name__ == "__main__":
    main()
